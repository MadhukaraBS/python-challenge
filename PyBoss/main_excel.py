import sys
import csv
from datetime import datetime
import openpyxl

# State name to abbreviation mapping table
us_state_abbrev = {
    'Alabama': 'AL',
    'Alaska': 'AK',
    'Arizona': 'AZ',
    'Arkansas': 'AR',
    'California': 'CA',
    'Colorado': 'CO',
    'Connecticut': 'CT',
    'Delaware': 'DE',
    'Florida': 'FL',
    'Georgia': 'GA',
    'Hawaii': 'HI',
    'Idaho': 'ID',
    'Illinois': 'IL',
    'Indiana': 'IN',
    'Iowa': 'IA',
    'Kansas': 'KS',
    'Kentucky': 'KY',
    'Louisiana': 'LA',
    'Maine': 'ME',
    'Maryland': 'MD',
    'Massachusetts': 'MA',
    'Michigan': 'MI',
    'Minnesota': 'MN',
    'Mississippi': 'MS',
    'Missouri': 'MO',
    'Montana': 'MT',
    'Nebraska': 'NE',
    'Nevada': 'NV',
    'New Hampshire': 'NH',
    'New Jersey': 'NJ',
    'New Mexico': 'NM',
    'New York': 'NY',
    'North Carolina': 'NC',
    'North Dakota': 'ND',
    'Ohio': 'OH',
    'Oklahoma': 'OK',
    'Oregon': 'OR',
    'Pennsylvania': 'PA',
    'Rhode Island': 'RI',
    'South Carolina': 'SC',
    'South Dakota': 'SD',
    'Tennessee': 'TN',
    'Texas': 'TX',
    'Utah': 'UT',
    'Vermont': 'VT',
    'Virginia': 'VA',
    'Washington': 'WA',
    'West Virginia': 'WV',
    'Wisconsin': 'WI',
    'Wyoming': 'WY',
}

#  Function to generate row ready to be written to CSV file
def gen_row(emp_id, name, fmt_date, ssn_part, state):
  name_part = name.strip().split()
#  print("{}, {}, {}, {}, ***-**-{}, {}".
#      format(emp_id, name_part[0], name_part[1], fmt_date.strftime('%m/%d/%y'),
#      ssn_part, us_state_abbrev[state]))
  return [emp_id, name_part[0], name_part[1], fmt_date.strftime('%m/%d/%y'),
      "***-**-" + ssn_part, us_state_abbrev[state]]



#  open data file and output file
#  Open XLSX file
wb = openpyxl.load_workbook('employee_data.xlsx')
wb_out = openpyxl.Workbook()
sheet = wb['employee_data']
std = wb_out['Sheet']
wb_out.remove(std)
sheet_out = wb_out.create_sheet('employee_data')
#  Following API is deprecated
#  sheet = wb.get_sheet_by_name('employee_data')

#  Write XLSX header only if the current file has atleast one input row
if sheet.max_row > 1:
  sheet_out.cell(row=1, column=1).value = "Emp ID"
  sheet_out.cell(row=1, column=2).value = "First Name"
  sheet_out.cell(row=1, column=3).value = "Last Name"
  sheet_out.cell(row=1, column=4).value = "DOB"
  sheet_out.cell(row=1, column=5).value = "SSN"
  sheet_out.cell(row=1, column=6).value = "State"

#  Initialize ssn list
ssn_list = []

#  Step thru each row of CSV file
for row in range(2, sheet.max_row + 1):
  emp_id = sheet['A' + str(row)].value
  name = sheet['B' + str(row)].value
  name_part = name.strip().split()
  dob = sheet['C' + str(row)].value
  ssn = sheet['D' + str(row)].value
  state = sheet['E' + str(row)].value
  ssn_list = ssn.split("-")
  state = state.strip()
  #  print(gen_row(emp_id, name, dob, ssn_list[2], state))
  sheet_out.cell(row=row, column=1).value = emp_id
  sheet_out.cell(row=row, column=2).value = name_part[0]
  sheet_out.cell(row=row, column=3).value = name_part[1]
  sheet_out.cell(row=row, column=4).value = dob.strftime('%m/%d/%Y')
  sheet_out.cell(row=row, column=5).value = "***-**-" + ssn_list[2]
  sheet_out.cell(row=row, column=6).value = us_state_abbrev[state]
wb_out.save('processed_xlsx_emp_data.xlsx')
